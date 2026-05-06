import openpyxl
file="C:\\Users\\Asus\\OneDrive\\Documents\\login.xlsx"
workbook=openpyxl.load_workbook(file)
sheet = workbook['Sheet3']
#write the data from web to xl
#for same data
#for r in range (1,6):
#    for c in range (1,6):
#        sheet.cell(r,c).value = "login"

#different data
sheet.cell(1,1).value = "no."
sheet.cell(1,2).value = "name"
sheet.cell(1,3).value = "dept"

sheet.cell(2,1).value = "1"
sheet.cell(2,2).value = "yash"
sheet.cell(2,3).value = "test"

sheet.cell(3,1).value = "2"
sheet.cell(3,2).value = "yashu"
sheet.cell(3,3).value = "dev."

workbook.save(file)


