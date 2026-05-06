from openpyxl import load_workbook

def getRowCount(file, sheetName):
    wb = load_workbook(file)
    sheet = wb[sheetName]
    rowCount = sheet.max_row
    wb.close()
    return rowCount

def readData(file, sheetName, rownum, columnno):
    wb = load_workbook(file)
    sheet = wb[sheetName]
    data = sheet.cell(row=rownum, column=columnno).value
    wb.close()
    return data

def deleteRow(file, sheetName, rowNum):
    wb = load_workbook(file)
    sheet = wb[sheetName]
    sheet.delete_rows(rowNum, 1)
    wb.save(file)
    wb.close()

def deleteRowAndColumn(file, sheet_name, row_index, column_index):
    wb = load_workbook(file)
    sheet = wb[sheet_name]
    sheet.delete_rows(row_index)
    sheet.delete_cols(column_index)
    wb.save(file)
    wb.close()